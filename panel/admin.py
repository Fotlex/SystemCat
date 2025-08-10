import openpyxl

from openpyxl.styles import Font
from io import BytesIO

from django.contrib import admin
from panel.models import *
from django.http import HttpResponse

from django.db.models import (
    Count, Avg, Sum, F, ExpressionWrapper, DurationField, Q, Value, DecimalField
)
from django.db.models.functions import Coalesce


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ('id', 'username', 'first_name', 'last_name', 'role')
    fields = ('id', 'username', 'first_name', 'last_name', 'role')

    exclude = ('data',)


class OrderItemInline(admin.TabularInline):
    model = OrderItem  
    
    fields = ('product_type', 'quantity', 'size', 'color', 'price',)
    
    extra = 1  
    



def auto_adjust_column_width(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_full_report_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws_analytics = wb.create_sheet("Аналитика")
    ws_order_details = wb.create_sheet("Детализация по заказам")
    ws_clients = wb.create_sheet("Клиенты")
    ws_users = wb.create_sheet("Сотрудники")

    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=14)

    current_row = 1

    ws_analytics.cell(row=current_row, column=1, value="Количество заказов по статусам").font = header_font
    current_row += 1
    ws_analytics.cell(row=current_row, column=1, value="Статус").font = bold_font
    ws_analytics.cell(row=current_row, column=2, value="Количество").font = bold_font
    current_row += 1
    status_counts = Order.objects.values('status').annotate(count=Count('id')).order_by('status')
    status_map = dict(Order.STATUS_CHOICES)
    for item in status_counts:
        ws_analytics.cell(row=current_row, column=1, value=status_map.get(item['status'], item['status']))
        ws_analytics.cell(row=current_row, column=2, value=item['count'])
        current_row += 1
    current_row += 2

    ws_analytics.cell(row=current_row, column=1, value="Среднее время выполнения заказа").font = header_font
    current_row += 1
    completed_orders = Order.objects.filter(status='completed', created_at__isnull=False, completed_at__isnull=False)
    avg_duration_result = completed_orders.aggregate(avg_duration=Avg(ExpressionWrapper(F('completed_at') - F('created_at'), output_field=DurationField())))
    avg_duration = avg_duration_result.get('avg_duration')
    if avg_duration:
        days = avg_duration.days
        seconds = avg_duration.seconds
        hours = seconds // 3600
        minutes = (seconds % 3600) // 60
        avg_time_str = f"{days} дн. {hours} ч. {minutes} мин."
    else:
        avg_time_str = "Нет завершенных заказов для расчета"
    ws_analytics.cell(row=current_row, column=1, value="Среднее время (от создания до завершения)").font = bold_font
    ws_analytics.cell(row=current_row, column=2, value=avg_time_str)
    current_row += 3

    ws_analytics.cell(row=current_row, column=1, value="Количество заказов по типам").font = header_font
    current_row += 1
    ws_analytics.cell(row=current_row, column=1, value="Тип заказа").font = bold_font
    ws_analytics.cell(row=current_row, column=2, value="Количество").font = bold_font
    current_row += 1
    type_counts = Order.objects.values('order_type').annotate(count=Count('id')).order_by('order_type')
    type_map = dict(Order.ORDER_TYPE_CHOICES)
    for item in type_counts:
        ws_analytics.cell(row=current_row, column=1, value=type_map.get(item['order_type'], item['order_type']))
        ws_analytics.cell(row=current_row, column=2, value=item['count'])
        current_row += 1
    current_row += 2

    ws_analytics.cell(row=current_row, column=1, value="Статистика по сотрудникам").font = header_font
    current_row += 1
    ws_analytics.cell(row=current_row, column=1, value="Сотрудник").font = bold_font
    ws_analytics.cell(row=current_row, column=2, value="Принял заказов (история)").font = bold_font # <-- Меняем заголовок
    ws_analytics.cell(row=current_row, column=3, value="Завершил заказов").font = bold_font
    current_row += 1

    accepted_counts = OrderAssignmentLog.objects.values('employee').annotate(
        count=Count('order', distinct=True)
    )
    accepted_map = {item['employee']: item['count'] for item in accepted_counts}

    employees = User.objects.filter(Q(orders__isnull=False) | Q(cans_orders__isnull=False) | Q(assignment_logs__isnull=False)).distinct()
    for emp in employees:
        accepted_count = accepted_map.get(emp.id, 0)
        

        completed_count = Order.objects.filter(responsible_employee=emp, status='completed').count()
        
        if accepted_count > 0 or completed_count > 0:
            ws_analytics.cell(row=current_row, column=1, value=str(emp))
            ws_analytics.cell(row=current_row, column=2, value=accepted_count)
            ws_analytics.cell(row=current_row, column=3, value=completed_count)
            current_row += 1
    current_row += 2


    ws_analytics.cell(row=current_row, column=1, value="Общая сумма стоимостей").font = header_font
    current_row += 1
    total_measurement_cost = Order.objects.aggregate(total=Coalesce(Sum('measurement_cost'), Value(0.0, output_field=DecimalField())))['total']
    total_item_cost = OrderItem.objects.aggregate(total=Coalesce(Sum(F('price') * F('quantity'), output_field=DecimalField()), Value(0.0, output_field=DecimalField())))['total']
    ws_analytics.cell(row=current_row, column=1, value="Общая стоимость замеров")
    ws_analytics.cell(row=current_row, column=2, value=total_measurement_cost)
    current_row += 1
    ws_analytics.cell(row=current_row, column=1, value="Общая стоимость изделий")
    ws_analytics.cell(row=current_row, column=2, value=total_item_cost)
    current_row += 3


    ws_analytics.cell(row=current_row, column=1, value="Статистика по оплатам (на основе поля 'Расчет')").font = header_font
    current_row += 1
    orders_with_payment_info = Order.objects.exclude(genral_cost_info__isnull=True).exclude(genral_cost_info__exact='')
    total_relevant_orders_count = orders_with_payment_info.count()
    advance_orders_count = orders_with_payment_info.exclude(genral_cost_info='100%').count()
    if total_relevant_orders_count > 0:
        advance_percentage = (advance_orders_count / total_relevant_orders_count) * 100
        percentage_str = f"{advance_percentage:.2f}% ({advance_orders_count} из {total_relevant_orders_count})"
    else:
        percentage_str = "Нет заказов с информацией об оплате"
    ws_analytics.cell(row=current_row, column=1, value="Процент заказов с авансом").font = bold_font
    ws_analytics.cell(row=current_row, column=2, value=percentage_str)
    current_row += 3

    ws_analytics.cell(row=current_row, column=1, value="Количество заказов по местоположению").font = header_font
    current_row += 1
    location_counts = Order.objects.filter(subtype__in=['city', 'intercity']).values('subtype').annotate(count=Count('id'))
    location_map = {'city': 'Город', 'intercity': 'Межгород'}
    for item in location_counts:
        ws_analytics.cell(row=current_row, column=1, value=location_map.get(item['subtype'], item['subtype']))
        ws_analytics.cell(row=current_row, column=2, value=item['count'])
        current_row += 1
    

    client_row_map = {}
    user_row_map = {}

    client_headers = ['ID Клиента', 'Имя', 'Номер телефона', 'Адрес']
    ws_clients.append(client_headers)
    for cell in ws_clients[1]: cell.font = bold_font
    for row_num, client in enumerate(Client.objects.all(), 2):
        ws_clients.append([client.id, client.name, client.phone_number, client.address])
        client_row_map[client.id] = row_num


    user_headers = ['ID Сотрудника', 'Юзернейм', 'Имя', 'Фамилия', 'Роль', 'Дата регистрации']
    ws_users.append(user_headers)
    for cell in ws_users[1]: cell.font = bold_font
    for row_num, user in enumerate(User.objects.all(), 2):
        ws_users.append([user.id, user.username, user.first_name, user.last_name, user.get_role_display(), user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else ''])
        user_row_map[user.id] = row_num
        

    details_headers = ['ID Заказа', 'Клиент', 'Статус', 'Ответственный', 'Стоимость замера', 'Стоимость изделий', 'Итого']
    ws_order_details.append(details_headers)
    for cell in ws_order_details[1]: cell.font = bold_font

    orders_with_costs = Order.objects.select_related('client', 'responsible_employee').annotate(
        item_cost_sum=Coalesce(
            Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            Value(0.0, output_field=DecimalField())
        )
    ).order_by('-id')

    for row_num, order in enumerate(orders_with_costs, 2):
        measurement_cost = order.measurement_cost or 0
        item_cost = order.item_cost_sum
        total_cost = measurement_cost + item_cost
        
        client_str = str(order.client) if order.client else "N/A"
        responsible_str = str(order.responsible_employee) if order.responsible_employee else "N/A"

        row_data = [
            order.id, client_str, order.get_status_display(), responsible_str,
            measurement_cost, item_cost, total_cost
        ]
        ws_order_details.append(row_data)

        if order.client and order.client.id in client_row_map:
            client_row_in_excel = client_row_map[order.client.id]
            client_cell = ws_order_details.cell(row=row_num, column=2)
            client_cell.hyperlink = f"#'Клиенты'!A{client_row_in_excel}"
            client_cell.style = "Hyperlink"
        
        if order.responsible_employee and order.responsible_employee.id in user_row_map:
            user_row_in_excel = user_row_map[order.responsible_employee.id]
            user_cell = ws_order_details.cell(row=row_num, column=4)
            user_cell.hyperlink = f"#'Сотрудники'!A{user_row_in_excel}"
            user_cell.style = "Hyperlink"

    auto_adjust_column_width(ws_analytics)
    auto_adjust_column_width(ws_order_details)
    auto_adjust_column_width(ws_clients)
    auto_adjust_column_width(ws_users)
    
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    response = HttpResponse(
        virtual_workbook.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="full_report.xlsx"'
    
    return response

export_full_report_to_excel.short_description = "Экспортировать полный отчет в Excel"



@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    search_fields = [
        'id', 
        'client__name', 
        'client__phone_number',
        'client__address',
    ]

    list_filter = ('status', 'order_type')
    readonly_fields = ('created_at', 'completed_at') 
    
    actions = [export_full_report_to_excel]
    
    inlines = [
        OrderItemInline,
    ]

admin.site.register(Client)
